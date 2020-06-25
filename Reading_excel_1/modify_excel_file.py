import openpyxl

# Program rerun will be have error because file "my_example.xlsx" is already available
# Firstly, program should be check whether file "my_example.xlsx" is available or not
# if it is available, delete it

wb = openpyxl.Workbook()
print(wb)
#wb.sheetnames

#sheet = wb.get_sheet_by_name('Sheet')
sheet = wb.active

sheet['A1'] = 56 
sheet['A2'] = 42

# Create a new sheet with name = sheet2
sheet2 = wb.create_sheet()
sheet2['A1'] = "Hello WOrld"
sheet2['A2'] = 12345678

# Create a new sheet with name = sheet3
sheet3 = wb.create_sheet(title = "mysheet")
sheet3['A1'] = "Chao the gioi"
sheet3['A2'] = "Toi la Binh day"
import os
os.chdir('F:\\python_execsice\\Reading_excel')
wb.save('my_example.xlsx')
